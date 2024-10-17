import cv2
import mediapipe as mp
import math
import numpy as np
from ctypes import cast, POINTER
from comtypes import CLSCTX_ALL
from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
import time
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Paths
volume_data_path = "D:\\Python Files\\MEDIBOT_FINAL\\Volume\\Image"
excel_path_vol = "D:\\Python Files\\MEDIBOT_FINAL\\Volume\\VolumeData.xlsx"


# Replace the URL with the IP camera's stream URL
url = 'http://192.168.1.132/cam-hi.jpg'

# Create directories if they don't exist
if not os.path.exists(volume_data_path):
    os.makedirs(volume_data_path)

def volmain():
    # Solution APIs
    mp_drawing = mp.solutions.drawing_utils
    mp_hands = mp.solutions.hands

    # Volume Control Setup using PyCaw
    devices = AudioUtilities.GetSpeakers()
    interface = devices.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
    volume = cast(interface, POINTER(IAudioEndpointVolume))
    volRange = volume.GetVolumeRange()
    minVol, maxVol = volRange[0], volRange[1]
    volBar, volPer = 400, 0
    volume_list = []  # List to store detected volume levels
    start_time = time.time()  # Time tracker for the duration of the program

    #cam = cv2.VideoCapture(0)
    cam = cv2.VideoCapture(url)
    if not cam.isOpened():
        print("Error: Could not open video stream from IP camera")
        return

    # Mediapipe Hand Landmark Model
    with mp_hands.Hands(
        model_complexity=0,
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5) as hands:

        frame_count = 0

        while cam.isOpened():
            success, image = cam.read()
            if not success:
                break

            # Update the current time for each frame
            current_time = time.time()

            # Convert the image to RGB for MediaPipe processing
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            results = hands.process(image)
            image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

            # Process detected hand landmarks
            if results.multi_hand_landmarks:
                for hand_landmarks in results.multi_hand_landmarks:
                    mp_drawing.draw_landmarks(
                        image,
                        hand_landmarks,
                        mp_hands.HAND_CONNECTIONS,
                        mp_drawing.DrawingSpec(color=(121, 22, 76), thickness=2, circle_radius=4),
                        mp_drawing.DrawingSpec(color=(250, 44, 250), thickness=2, circle_radius=2),
                    )

                # Extract landmarks for volume control logic
                lmList = []
                myHand = results.multi_hand_landmarks[0]  # Using the first hand detected
                for id, lm in enumerate(myHand.landmark):
                    h, w, _ = image.shape
                    cx, cy = int(lm.x * w), int(lm.y * h)
                    lmList.append([id, cx, cy])

                # Check if thumb and index finger landmarks are detected
                if lmList:
                    x1, y1 = lmList[4][1], lmList[4][2]  # Thumb tip
                    x2, y2 = lmList[8][1], lmList[8][2]  # Index finger tip

                    # Draw a line between thumb and index finger
                    cv2.line(image, (x1, y1), (x2, y2), (0, 255, 0), 3)

                    # Calculate distance between thumb and index finger
                    length = math.hypot(x2 - x1, y2 - y1)

                    # Map the distance to volume range
                    vol = np.interp(length, [50, 220], [minVol, maxVol])
                    volBar = np.interp(length, [50, 220], [400, 150])
                    volPer = np.interp(length, [50, 220], [0, 100])

                    # Set volume
                    volume.SetMasterVolumeLevel(vol, None)

                    # Add the detected volume percentage to the list
                    if current_time - start_time <= 60:  # Limit duration to 60 seconds
                        frame_count += 1
                        image_filename = os.path.join(volume_data_path, f"frame_{frame_count}.jpg")
                        cv2.imwrite(image_filename, image)  # Save frame

                        volume_list.append({
                            'Volume Percentage': int(volPer),
                            'Image Path': image_filename
                        })

                    # Display volume bar and percentage
                    cv2.rectangle(image, (50, 150), (85, 400), (0, 255, 0), 3)
                    cv2.rectangle(image, (50, int(volBar)), (85, 400), (0, 255, 0), cv2.FILLED)
                    cv2.putText(image, f'{int(volPer)} %', (40, 450), cv2.FONT_HERSHEY_COMPLEX,
                                1, (0, 255, 0), 3)

            # Show the video feed with hand landmarks and volume bar
            cv2.imshow('Hand Volume Control', image)

            # Exit loop after 60 seconds
            if current_time - start_time > 60:
                break

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

    # After the loop ends, save the volume data to Excel
    cam.release()
    cv2.destroyAllWindows()

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # Create a DataFrame for the collected data
    data = []
    count_less_50 = 0
    count_greater_50 = 0

    for vol_data in volume_list:
        data.append({
            'Date': date_str,
            'Time': time_str,
            'Volume Percentage': vol_data['Volume Percentage'],
            'Image Path': vol_data['Image Path']
        })
        if vol_data['Volume Percentage'] < 50:
            count_less_50 += 1
        else:
            count_greater_50 += 1

    df = pd.DataFrame(data)

    # Save or append data to the existing Excel file
    try:
        sheet_name = 'Volume Data'
        if os.path.exists(excel_path_vol):
            # Load existing workbook
            book = load_workbook(excel_path_vol)
            with pd.ExcelWriter(excel_path_vol, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer.book = book
                if sheet_name not in writer.book.sheetnames:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=writer.sheets[sheet_name].max_row)
        else:
            # If file doesn't exist, create new Excel file
            with pd.ExcelWriter(excel_path_vol, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    except PermissionError as e:
        print(f"Permission Error: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

    # Display the volume data and counts
    print(df)
    print(f"Count of Volume < 50 (Danger): {count_less_50}")
    print(f"Count of Volume >= 50 (Normal): {count_greater_50}")

if __name__ == "__main__":
    volmain()
